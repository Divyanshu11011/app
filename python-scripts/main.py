import json
import pytesseract
from PIL import Image
import ftfy
import os
import io
import sys
import re
import openpyxl
from pan_read import pan_read_data
from aadhaar_read import adhaar_read_data
import mysql.connector
import mysql
# Set the correct path to the 'uploads' directory
uploads_dir = 'uploads'

# Get the current script directory
script_dir = os.path.dirname(os.path.abspath(__file__))

# Construct the correct path to the 'uploads' directory
uploads_full_path = os.path.join(script_dir, uploads_dir)

# Rename existing files in ascending order
list_of_files = os.listdir(uploads_full_path)
sorted_files = sorted(
    list_of_files,
    key=lambda x: int(re.search(r'\d+', x).group()) if re.search(r'\d+', x) else float('inf')
)

for idx, old_file_name in enumerate(sorted_files, start=1):
    old_file_path = os.path.join(uploads_full_path, old_file_name)

    # Generate a new filename
    new_file_name = f"{idx}{os.path.splitext(old_file_name)[1]}"
    new_file_path = os.path.join(uploads_full_path, new_file_name)

    # Check if the new filename already exists and increment index if needed
    while os.path.exists(new_file_path):
        idx += 1
        new_file_name = f"{idx}{os.path.splitext(old_file_name)[1]}"
        new_file_path = os.path.join(uploads_full_path, new_file_name)

    os.rename(old_file_path, new_file_path)

# Updated code to get the latest file in the 'uploads' directory
list_of_files = os.listdir(uploads_full_path)

# Check if the directory is empty
if not list_of_files:
    print("No files found in the 'uploads' directory.")
    sys.exit(1)

# Get the latest file with full path
latest_file = sorted(list_of_files)[-1]
latest_file_path = os.path.join(uploads_full_path, latest_file)

img = Image.open(latest_file_path)  # Use the latest file

# Text preprocessing and data extraction
text = pytesseract.image_to_string(img, lang='eng')

# Classify data and generate JSON output
if any(keyword in text.lower() for keyword in ["income", "tax", "department"]):
    data = pan_read_data(text)
    sheet_name = 'Pan Details'
    headers = ['Date of Birth', 'Father Name', 'ID Type', 'Name', 'PAN']
elif "male" in text.lower():
    data = adhaar_read_data(text)
    sheet_name = 'Aadhar Details'
    headers = ['Adhaar Number', 'Date of Birth', 'ID Type', 'Name', 'Sex']
else:
    data = {'ID TYPE': 'Unrecognized'}
    sheet_name = 'Unknown Details'
    headers = []

json_output_path = 'info.json'
with io.open(json_output_path, 'w', encoding='utf-8') as outfile:
    data_json = json.dumps(data, indent=4, sort_keys=True, separators=(',', ': '), ensure_ascii=False)
    outfile.write(data_json)

# Reading JSON data
with open(json_output_path, encoding='utf-8') as data_file:
    data_loaded = json.load(data_file)

# Open or create an Excel workbook
excel_file_path = 'output_data.xlsx'
if os.path.exists(excel_file_path):
    workbook = openpyxl.load_workbook(excel_file_path)
else:
    workbook = openpyxl.Workbook()

# Check if the sheet already exists
if sheet_name not in workbook.sheetnames:
    workbook.create_sheet(sheet_name)
    
# Get the active sheet
active_sheet = workbook[sheet_name]

# Append headers if not already present
if not active_sheet['A1'].value:
    for col_num, header in enumerate(headers, start=1):
        active_sheet.cell(row=1, column=col_num, value=header)

# Append data to the Excel sheet
new_row = [data_loaded.get(header, '') for header in headers]
active_sheet.append(new_row)

# Save the workbook
workbook.save(excel_file_path)

db_config = {
    'host': '127.0.0.1',
    'user': 'root',
    'password': 'div123',
    'database': 'customeragentoutput',
}
#Establish connection with Mysql
try:
    connection = mysql.connector.connect(**db_config)
    cursor = connection.cursor()
except mysql.connector.Error as err:
    print(f"Error: {err}")
    sys.exit(1)

# Create Pan Details and Aadhaar Details tables if they don't exist
create_pan_table_query = """
CREATE TABLE IF NOT EXISTS pan_details (
    id INT AUTO_INCREMENT PRIMARY KEY,
    date_of_birth VARCHAR(255),
    father_name VARCHAR(255),
    id_type VARCHAR(255),
    name VARCHAR(255),
    pan VARCHAR(255)
)
"""

create_aadhaar_table_query = """
CREATE TABLE IF NOT EXISTS aadhaar_details (
    id INT AUTO_INCREMENT PRIMARY KEY,
    adhaar_number VARCHAR(255),
    date_of_birth VARCHAR(255),
    id_type VARCHAR(255),
    name VARCHAR(255),
    sex VARCHAR(255)
)
"""

try:
    cursor.execute(create_pan_table_query)
    cursor.execute(create_aadhaar_table_query)
    connection.commit()
except mysql.connector.Error as err:
    print(f"Error creating tables: {err}")
    connection.rollback()
    sys.exit(1)

if data_loaded.get('ID Type') == 'PAN':
    insert_pan_data_query = """
    INSERT INTO pan_details (date_of_birth, father_name, id_type, name, pan)
    VALUES (%s, %s, %s, %s, %s)
    """
    pan_data = (
        data_loaded.get('Date of Birth'),
        data_loaded.get('Father Name'),
        data_loaded.get('ID Type'),
        data_loaded.get('Name'),
        data_loaded.get('PAN'),
    )

    try:
        cursor.execute(insert_pan_data_query, pan_data)
        connection.commit()
    except mysql.connector.Error as err:
        print(f"Error inserting PAN data: {err}")
        connection.rollback()
else:
    # Assuming ID Type is Aadhaar
    insert_aadhaar_data_query = """
    INSERT INTO aadhaar_details (adhaar_number, date_of_birth, id_type, name, sex)
    VALUES (%s, %s, %s, %s, %s)
    """
    aadhaar_data = (
        data_loaded.get('Adhaar Number'),
        data_loaded.get('Date of Birth'),
        data_loaded.get('ID Type'),
        data_loaded.get('Name'),
        data_loaded.get('Sex'),
    )

    try:
        cursor.execute(insert_aadhaar_data_query, aadhaar_data)
        connection.commit()
    except mysql.connector.Error as err:
        print(f"Error inserting Aadhaar data: {err}")
        connection.rollback()


# Printing the data and cleaning up
if data_loaded.get('ID Type') == 'PAN':
    print("\n---------- PAN Details ----------")
    print("\nDate of Birth: ", data_loaded.get('Date of Birth'))
    print("\nFather Name: ", data_loaded.get('Father Name'))
    print("\nID Type: ", data_loaded.get('ID Type'))
    print("\nName: ", data_loaded.get('Name'))
    print("\nPAN: ", data_loaded.get('PAN'))
    print("\n---------------------------------")
elif data_loaded.get('ID Type') == 'Aadhaar':
    print("\n---------- Aadhaar Details ----------")
    print("\nAdhaar Number: ", data_loaded.get('Adhaar Number'))
    print("\nDate of Birth: ", data_loaded.get('Date of Birth'))
    print("\nID Type: ", data_loaded.get('ID Type'))
    print("\nName: ", data_loaded.get('Name'))
    print("\nSex: ", data_loaded.get('Sex'))
    print("\n------------------------------------")
else:
    print("Document type not recognized.")

# Remove the processed file
os.remove(latest_file_path)
cursor.close()
connection.close()